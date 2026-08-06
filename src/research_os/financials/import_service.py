"""财务导入服务（Phase 4 任务书 3.11/Commit 3）。

支持 CSV / JSON / XLSX 人工导入 → FinancialDataManifest + FinancialReport + FinancialFact。
规则：
- 所有输入建立文件哈希、数据版本、导入时间、导入人、来源类别、公司、报告期、口径、币种、单位；
- 行接受/拒绝状态、错误和警告显式记录；rejected 行不写正式事实表；
- dry-run 零副作用（不写库、不写 manifest）；
- checksum + data_version 进幂等键；
- 财务值一律十进制字符串，不得以 float 持久化。
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from research_os.models.financials import (
    FinancialDataManifest,
    FinancialFact,
    FinancialReport,
)
from research_os.utils.decimal import normalize_decimal_string
from research_os.utils.time import now_iso

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

# 必需列（CSV/XLSX 表头或 JSON 对象键）
REQUIRED_COLUMNS = [
    "company_entity_id",
    "period_start",
    "period_end",
    "fiscal_year",
    "report_type",
    "statement_scope",
    "statement_type",
    "taxonomy_code",
    "label_raw",
    "value",
    "unit_scale",
    "currency",
]

REPORT_TYPES = {"annual": "FY", "interim": "H1", "q1": "Q1", "q3": "Q3", "other": "OTHER"}
SCOPES = {"consolidated", "parent"}
STATEMENT_TYPES = {
    "income_statement", "balance_sheet", "cash_flow",
    "equity_statement", "note", "operating_data",
}


@dataclass
class RowResult:
    """单行导入结果。"""
    row_index: int
    accepted: bool
    issues: List[str] = field(default_factory=list)
    fact: Optional[FinancialFact] = None


@dataclass
class ImportResult:
    """一次导入的结果（含 manifest 与行级明细）。"""
    manifest: FinancialDataManifest
    rows: List[RowResult] = field(default_factory=list)
    reports: List[FinancialReport] = field(default_factory=list)

    @property
    def accepted_count(self) -> int:
        return sum(1 for r in self.rows if r.accepted)

    @property
    def rejected_count(self) -> int:
        return sum(1 for r in self.rows if not r.accepted)

    @property
    def errors(self) -> List[str]:
        return [f"行 {r.row_index}: {'; '.join(r.issues)}" for r in self.rows if not r.accepted]


def sha256_file(path: Path) -> str:
    """计算文件 SHA-256（确定性代码，指南 6.3）。"""
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _parse_value(value: Any) -> Optional[str]:
    """规范化数值为十进制字符串；空/NA 返回 None；非法数字抛 ValueError。"""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.upper() in ("NA", "N/A", "NULL", "NONE", "-"):
        return None
    try:
        return normalize_decimal_string(s)
    except ValueError as exc:
        raise ValueError(f"非法数字: {value!r}") from exc


def _parse_date(value: Any, field_name: str) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    if not DATE_RE.match(s):
        raise ValueError(f"{field_name} 必须是 YYYY-MM-DD: {value!r}")
    return s


def _default_disclosure_time(period_end: str) -> str:
    """财报披露惯例时间（保守估计，非导入时刻）：
    年报次年 4-30；中报当年 8-31；一季报当年 4-30；三季报当年 10-31。"""
    y = int(period_end[:4])
    if period_end.endswith("12-31"):
        return f"{y + 1}-04-30T00:00:00"
    if period_end.endswith("06-30"):
        return f"{y}-08-31T00:00:00"
    if period_end.endswith("03-31"):
        return f"{y}-04-30T00:00:00"
    if period_end.endswith("09-30"):
        return f"{y}-10-31T00:00:00"
    return f"{y}-12-31T00:00:00"


def _parse_disclosure(value: Any) -> Optional[str]:
    """解析真实披露时间（可选列，YYYY-MM-DD 或 ISO-8601）。"""
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    if DATE_RE.match(s):
        return f"{s}T00:00:00"
    if re.match(r"^\d{4}-\d{2}-\d{2}T", s):
        return s
    return None


def _row_to_dict(row: Dict[str, Any], col_map: Dict[str, str]) -> Dict[str, Any]:
    """按列映射（支持中文表头）提取标准键。"""
    out: Dict[str, Any] = {}
    for std_key in REQUIRED_COLUMNS:
        out[std_key] = row.get(std_key, row.get(col_map.get(std_key, ""), ""))
    return out


def _validate_row(data: Dict[str, Any], row_index: int, default_scope: str) -> RowResult:
    """校验并构造一行 FinancialFact。"""
    rr = RowResult(row_index=row_index, accepted=False)
    issues = rr.issues

    company = str(data.get("company_entity_id") or "").strip()
    if not company.startswith("company:"):
        issues.append("company_entity_id 必须以 company: 开头")

    period_end = _parse_date(data.get("period_end"), "period_end")
    if period_end is None:
        issues.append("period_end 缺失")

    report_type = str(data.get("report_type") or "").strip()
    if report_type not in REPORT_TYPES:
        issues.append(f"report_type 非法: {report_type!r}")

    scope = str(data.get("statement_scope") or default_scope or "").strip()
    if scope not in SCOPES:
        issues.append(f"statement_scope 非法: {scope!r}")

    stmt_type = str(data.get("statement_type") or "").strip()
    if stmt_type not in STATEMENT_TYPES:
        issues.append(f"statement_type 非法: {stmt_type!r}")

    taxonomy_code = str(data.get("taxonomy_code") or "").strip()
    if not taxonomy_code:
        issues.append("taxonomy_code 缺失")

    label_raw = str(data.get("label_raw") or "").strip()
    if not label_raw:
        issues.append("label_raw 缺失")

    currency = str(data.get("currency") or "").strip()
    if not re.match(r"^[A-Z]{3}$", currency):
        issues.append(f"currency 非法: {currency!r}")

    try:
        unit_scale = int(data.get("unit_scale") or 1)
        if unit_scale <= 0:
            raise ValueError
    except (TypeError, ValueError):
        unit_scale = 1
        issues.append("unit_scale 非法，回退 1")

    try:
        raw_value = _parse_value(data.get("value"))
    except ValueError as exc:
        raw_value = None
        issues.append(str(exc))

    if issues:
        return rr

    # 构造 FinancialFact（值一律十进制字符串；valid_from=报告期末，保证 <= as_of）
    fact = FinancialFact(
        fact_id=str(uuid.uuid4()),
        fact_key=f"{taxonomy_code}|{period_end}|{REPORT_TYPES[report_type]}|{scope}",
        financial_report_id="",
        company_entity_id=company,
        statement_type=stmt_type,
        taxonomy_code=taxonomy_code,
        label_raw=label_raw,
        period_start=_parse_date(data.get("period_start"), "period_start"),
        period_end=period_end,  # type: ignore[arg-type]
        instant_or_duration="duration" if stmt_type != "balance_sheet" else "instant",
        period_basis="reported_period",
        statement_scope=scope,  # type: ignore[arg-type]
        currency=currency,
        unit_scale=unit_scale,
        raw_value=raw_value,
        normalized_value=raw_value,
        normalized_unit="yuan",
        value_status="missing" if raw_value is None else "reported",
        sign_convention="reported",
        audit_status="unknown",
        segment_id=None,
        source_document_id=None,
        source_block_ids=[],
        evidence_ids=[],
        source_priority=5,
        restatement_version=1,
        valid_from=f"{period_end}T00:00:00",  # 事实自报告期末有效（<= as_of，ERV-053）
        valid_to=None,
        conflict_group_id=None,
        warnings=[],
        created_at=now_iso(),
    )
    # 真实披露时间（可选列，供 Evidence/未来信息检查使用；非模型字段）
    fact._published_at = _parse_disclosure(data.get("published_at"))
    rr.fact = fact
    rr.accepted = True
    return rr


def _read_csv(path: Path, col_map: Dict[str, str]) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        return [dict(r) for r in reader]


def _read_json(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = data.get("rows", [])
    if not isinstance(data, list):
        raise ValueError("JSON 必须是对象数组或 {rows: [...]}")
    return [dict(r) for r in data]


def _read_xlsx(path: Path, col_map: Dict[str, str]) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    result: List[Dict[str, Any]] = []
    for values in rows_iter:
        result.append({header[i]: v for i, v in enumerate(values) if i < len(header)})
    wb.close()
    return result


def import_financial_file(
    path: Path,
    *,
    company_entity_id: str,
    source_id: str = "manual_financial_import",
    imported_by: str = "user",
    data_version: str = "v1",
    default_scope: str = "consolidated",
    dry_run: bool = False,
) -> ImportResult:
    """导入一个财务文件，返回 ImportResult（dry_run 时仍解析但不落库）。"""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv(path, {})
        file_format = "csv"
    elif suffix == ".json":
        rows = _read_json(path)
        file_format = "json"
    elif suffix in (".xlsx", ".xlsm"):
        rows = _read_xlsx(path, {})
        file_format = "xlsx"
    else:
        raise ValueError(f"不支持的格式: {suffix}（仅 csv/json/xlsx）")

    checksum = sha256_file(path)
    imported_at = now_iso()

    # 行级校验
    results = [
        _validate_row(r, idx + 1, default_scope) for idx, r in enumerate(rows)
    ]

    manifest = FinancialDataManifest(
        manifest_id=str(uuid.uuid4()),
        source_kind="manual_import",
        source_id=source_id,
        file_name=path.name,
        file_format=file_format,  # type: ignore[arg-type]
        file_checksum=checksum,
        imported_at=imported_at,
        imported_by=imported_by,
        company_entity_ids=[company_entity_id],
        document_ids=[],
        report_period_start=None,
        report_period_end=None,
        default_statement_scope=default_scope,  # type: ignore[arg-type]
        default_currency=None,
        default_unit_scale=None,
        row_count=len(rows),
        accepted_count=0,
        rejected_count=0,
        data_version=data_version,
        validation_status="pending",
        validation_errors=[],
        warnings=[],
        source_ids=[source_id],
        version=1,
    )
    # 用实际计数覆盖
    manifest.row_count = len(rows)
    manifest.accepted_count = sum(1 for r in results if r.accepted)
    manifest.rejected_count = sum(1 for r in results if not r.accepted)
    manifest.validation_status = (
        "accepted" if manifest.rejected_count == 0 else
        ("partial" if manifest.accepted_count > 0 else "rejected")
    )
    if manifest.rejected_count:
        manifest.validation_errors = [f"行 {r.row_index}: {'; '.join(r.issues)}" for r in results if not r.accepted]

    result = ImportResult(manifest=manifest, rows=results)

    # 构造 FinancialReport 对象（按期间聚合；dry-run 同样构造但不落库）
    # 真实披露时间：文件提供 published_at 列则用之；否则默认财报发布惯例
    # （年报次年 4 月底前、中报当年 8 月底前）；绝不以导入时刻冒充披露时间。
    period_keys: Dict[tuple, FinancialReport] = {}
    period_fact_map: Dict[tuple, List[FinancialFact]] = {}
    period_published: Dict[tuple, str] = {}
    for rr in results:
        if not rr.accepted or rr.fact is None:
            continue
        f = rr.fact
        key = (f.period_end, f.statement_scope)
        period_fact_map.setdefault(key, []).append(f)
        if key in period_keys:
            continue
        period_published[key] = f._published_at or _default_disclosure_time(f.period_end)
        period_keys[key] = FinancialReport(
                financial_report_id=str(uuid.uuid4()),
                company_entity_id=company_entity_id,
                document_id=None,
                manifest_id=manifest.manifest_id,
                report_type=(
                    "annual" if f.period_end.endswith("12-31") else
                    "interim" if f.period_end.endswith("06-30") else
                    "q1" if f.period_end.endswith("03-31") else
                    "q3" if f.period_end.endswith("09-30") else "other"
                ),
                period_start=f.period_start,
                period_end=f.period_end,
                fiscal_year=int(f.period_end[:4]),
                fiscal_period=REPORT_TYPES.get(  # type: ignore[arg-type]
                    "annual" if f.period_end.endswith("12-31") else
                    "interim" if f.period_end.endswith("06-30") else
                    "q1" if f.period_end.endswith("03-31") else
                    "q3" if f.period_end.endswith("09-30") else "other",
                    "OTHER",
                ),
                duration_months=12 if f.period_end.endswith("12-31") else 6,
                statement_scope=f.statement_scope,
                accounting_standard="CAS",
                currency=f.currency,
                unit_scale=f.unit_scale,
                audit_status="unknown",
                audit_opinion="unknown",
                restatement_status="original",
                supersedes_report_id=None,
                filing_version=data_version,
                source_ids=[source_id],
                evidence_ids=[],
                data_status="complete",
                version=1,
                published_at=period_published[key],
                created_at=imported_at,
            )
        result.reports.append(period_keys[key])
    # 事实关联到报告对象
    for key, facts in period_fact_map.items():
        report = period_keys[key]
        for f in facts:
            f.financial_report_id = report.financial_report_id
    return result


def persist_import(db: Any, result: ImportResult) -> None:
    """将导入结果写入数据库（幂等：同 checksum+data_version 不重复写入）。

    db: 提供 upsert(obj)/query(sql, params) 的 Database 实例。调用方负责 Schema 校验。
    """
    if result.manifest.rejected_count > 0 and result.manifest.accepted_count == 0:
        raise ValueError("全部行被拒绝，不写入正式表")
    # 幂等命中：同 checksum+data_version 的 manifest 已存在 → 跳过写入
    existing = db.query(
        "SELECT manifest_id FROM financial_data_manifests WHERE file_checksum = ? AND data_version = ?",
        (result.manifest.file_checksum, result.manifest.data_version),
    )
    if existing:
        return
    db.upsert(result.manifest)
    for report in result.reports:
        db.upsert(report)
    for rr in result.rows:
        if rr.accepted and rr.fact is not None:
            db.upsert(rr.fact)
